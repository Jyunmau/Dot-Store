"""
Dot-Store V2.1 报表服务层
"""
from datetime import datetime, timedelta, date
from decimal import Decimal
from typing import Optional, Dict, List, Any
from calendar import monthrange
from io import BytesIO

from sqlalchemy.orm import Session
from sqlalchemy import and_, func, extract

from app.models.transaction import Transaction
from app.models.order import Order


class ReportService:
    """
    报表服务类
    """

    def __init__(self, db: Session):
        """
        初始化报表服务
        """
        self.db = db

    def get_daily_report(self, user_id: int, report_date: Optional[date] = None) -> dict:
        """
        获取每日报表
        
        Args:
            user_id: 用户ID
            report_date: 报表日期，默认为今天
            
        Returns:
            dict: 包含date, income, expense, profit, categories的字典
        """
        if report_date is None:
            report_date = date.today()
        
        start_datetime = datetime.combine(report_date, datetime.min.time())
        end_datetime = datetime.combine(report_date, datetime.max.time())
        
        transactions = self.db.query(Transaction).filter(
            and_(
                Transaction.user_id == user_id,
                Transaction.created_at >= start_datetime,
                Transaction.created_at <= end_datetime
            )
        ).all()
        
        orders = self.db.query(Order).filter(
            and_(
                Order.user_id == user_id,
                Order.created_at >= start_datetime,
                Order.created_at <= end_datetime,
                Order.is_deleted == False
            )
        ).all()
        
        income = Decimal("0")
        expense = Decimal("0")
        income_categories: Dict[str, Decimal] = {}
        expense_categories: Dict[str, Decimal] = {}
        
        for t in transactions:
            if t.type == "income":
                income += t.amount
                if t.category in income_categories:
                    income_categories[t.category] += t.amount
                else:
                    income_categories[t.category] = t.amount
            elif t.type == "expense":
                expense += t.amount
                if t.category in expense_categories:
                    expense_categories[t.category] += t.amount
                else:
                    expense_categories[t.category] = t.amount
        
        order_count = len(orders)
        order_amount = sum(o.amount for o in orders)
        
        profit = income - expense
        
        return {
            "date": report_date.isoformat(),
            "income": float(income),
            "expense": float(expense),
            "profit": float(profit),
            "order_count": order_count,
            "order_amount": float(order_amount),
            "income_categories": {k: float(v) for k, v in income_categories.items()},
            "expense_categories": {k: float(v) for k, v in expense_categories.items()},
        }

    def get_weekly_report(self, user_id: int, start_date: Optional[date] = None, end_date: Optional[date] = None) -> dict:
        """
        获取每周报表
        
        Args:
            user_id: 用户ID
            start_date: 开始日期，默认为本周一
            end_date: 结束日期，默认为本周日
            
        Returns:
            dict: 包含start_date, end_date, income, expense, profit, daily_data, categories的字典
        """
        today = date.today()
        if start_date is None:
            start_date = today - timedelta(days=today.weekday())
        if end_date is None:
            end_date = start_date + timedelta(days=6)
        
        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())
        
        transactions = self.db.query(Transaction).filter(
            and_(
                Transaction.user_id == user_id,
                Transaction.created_at >= start_datetime,
                Transaction.created_at <= end_datetime
            )
        ).all()
        
        orders = self.db.query(Order).filter(
            and_(
                Order.user_id == user_id,
                Order.created_at >= start_datetime,
                Order.created_at <= end_datetime,
                Order.is_deleted == False
            )
        ).all()
        
        income = Decimal("0")
        expense = Decimal("0")
        income_categories: Dict[str, Decimal] = {}
        expense_categories: Dict[str, Decimal] = {}
        daily_data: List[Dict[str, Any]] = []
        
        daily_stats: Dict[date, Dict[str, Decimal]] = {}
        current_date = start_date
        while current_date <= end_date:
            daily_stats[current_date] = {"income": Decimal("0"), "expense": Decimal("0"), "profit": Decimal("0")}
            current_date += timedelta(days=1)
        
        for t in transactions:
            t_date = t.created_at.date()
            if t_date in daily_stats:
                if t.type == "income":
                    income += t.amount
                    daily_stats[t_date]["income"] += t.amount
                    if t.category in income_categories:
                        income_categories[t.category] += t.amount
                    else:
                        income_categories[t.category] = t.amount
                elif t.type == "expense":
                    expense += t.amount
                    daily_stats[t_date]["expense"] += t.amount
                    if t.category in expense_categories:
                        expense_categories[t.category] += t.amount
                    else:
                        expense_categories[t.category] = t.amount
        
        for d, stats in daily_stats.items():
            stats["profit"] = stats["income"] - stats["expense"]
            daily_data.append({
                "date": d.isoformat(),
                "income": float(stats["income"]),
                "expense": float(stats["expense"]),
                "profit": float(stats["profit"])
            })
        
        order_count = len(orders)
        order_amount = sum(o.amount for o in orders)
        profit = income - expense
        
        return {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "income": float(income),
            "expense": float(expense),
            "profit": float(profit),
            "order_count": order_count,
            "order_amount": float(order_amount),
            "daily_data": daily_data,
            "income_categories": {k: float(v) for k, v in income_categories.items()},
            "expense_categories": {k: float(v) for k, v in expense_categories.items()},
        }

    def get_monthly_report(self, user_id: int, year: Optional[int] = None, month: Optional[int] = None) -> dict:
        """
        获取每月报表
        
        Args:
            user_id: 用户ID
            year: 年份，默认为当前年
            month: 月份，默认为当前月
            
        Returns:
            dict: 包含year, month, income, expense, profit, weekly_data, categories的字典
        """
        today = date.today()
        if year is None:
            year = today.year
        if month is None:
            month = today.month
        
        _, days_in_month = monthrange(year, month)
        start_date = date(year, month, 1)
        end_date = date(year, month, days_in_month)
        
        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())
        
        transactions = self.db.query(Transaction).filter(
            and_(
                Transaction.user_id == user_id,
                Transaction.created_at >= start_datetime,
                Transaction.created_at <= end_datetime
            )
        ).all()
        
        orders = self.db.query(Order).filter(
            and_(
                Order.user_id == user_id,
                Order.created_at >= start_datetime,
                Order.created_at <= end_datetime,
                Order.is_deleted == False
            )
        ).all()
        
        income = Decimal("0")
        expense = Decimal("0")
        income_categories: Dict[str, Decimal] = {}
        expense_categories: Dict[str, Decimal] = {}
        weekly_data: List[Dict[str, Any]] = []
        
        weekly_stats: Dict[int, Dict[str, Decimal]] = {}
        for week in range(1, 6):
            weekly_stats[week] = {"income": Decimal("0"), "expense": Decimal("0"), "profit": Decimal("0")}
        
        for t in transactions:
            t_date = t.created_at.date()
            week_num = (t_date.day - 1) // 7 + 1
            if week_num > 5:
                week_num = 5
            
            if t.type == "income":
                income += t.amount
                weekly_stats[week_num]["income"] += t.amount
                if t.category in income_categories:
                    income_categories[t.category] += t.amount
                else:
                    income_categories[t.category] = t.amount
            elif t.type == "expense":
                expense += t.amount
                weekly_stats[week_num]["expense"] += t.amount
                if t.category in expense_categories:
                    expense_categories[t.category] += t.amount
                else:
                    expense_categories[t.category] = t.amount
        
        for week, stats in weekly_stats.items():
            stats["profit"] = stats["income"] - stats["expense"]
            weekly_data.append({
                "week": week,
                "income": float(stats["income"]),
                "expense": float(stats["expense"]),
                "profit": float(stats["profit"])
            })
        
        order_count = len(orders)
        order_amount = sum(o.amount for o in orders)
        profit = income - expense
        
        return {
            "year": year,
            "month": month,
            "income": float(income),
            "expense": float(expense),
            "profit": float(profit),
            "order_count": order_count,
            "order_amount": float(order_amount),
            "weekly_data": weekly_data,
            "income_categories": {k: float(v) for k, v in income_categories.items()},
            "expense_categories": {k: float(v) for k, v in expense_categories.items()},
        }

    def get_custom_report(
        self,
        user_id: int,
        start_date: date,
        end_date: date,
        transaction_type: Optional[str] = None,
        categories: Optional[List[str]] = None
    ) -> dict:
        """
        获取自定义报表
        
        Args:
            user_id: 用户ID
            start_date: 开始日期
            end_date: 结束日期
            transaction_type: 类型筛选（income/expense/all）
            categories: 分类筛选列表
            
        Returns:
            dict: 包含start_date, end_date, income, expense, profit, categories的字典
        """
        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())
        
        query = self.db.query(Transaction).filter(
            and_(
                Transaction.user_id == user_id,
                Transaction.created_at >= start_datetime,
                Transaction.created_at <= end_datetime
            )
        )
        
        if transaction_type and transaction_type != "all":
            query = query.filter(Transaction.type == transaction_type)
        
        if categories:
            query = query.filter(Transaction.category.in_(categories))
        
        transactions = query.all()
        
        orders_query = self.db.query(Order).filter(
            and_(
                Order.user_id == user_id,
                Order.created_at >= start_datetime,
                Order.created_at <= end_datetime,
                Order.is_deleted == False
            )
        )
        orders = orders_query.all()
        
        income = Decimal("0")
        expense = Decimal("0")
        income_categories: Dict[str, Decimal] = {}
        expense_categories: Dict[str, Decimal] = {}
        
        for t in transactions:
            if t.type == "income":
                income += t.amount
                if t.category in income_categories:
                    income_categories[t.category] += t.amount
                else:
                    income_categories[t.category] = t.amount
            elif t.type == "expense":
                expense += t.amount
                if t.category in expense_categories:
                    expense_categories[t.category] += t.amount
                else:
                    expense_categories[t.category] = t.amount
        
        order_count = len(orders)
        order_amount = sum(o.amount for o in orders)
        profit = income - expense
        
        return {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "income": float(income),
            "expense": float(expense),
            "profit": float(profit),
            "order_count": order_count,
            "order_amount": float(order_amount),
            "income_categories": {k: float(v) for k, v in income_categories.items()},
            "expense_categories": {k: float(v) for k, v in expense_categories.items()},
        }

    def get_category_analysis(
        self,
        user_id: int,
        start_date: date,
        end_date: date,
        transaction_type: str = "all"
    ) -> dict:
        """
        获取分类分析
        
        Args:
            user_id: 用户ID
            start_date: 开始日期
            end_date: 结束日期
            transaction_type: 类型（income/expense/all）
            
        Returns:
            dict: 包含type, categories, total的字典
        """
        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())
        
        query = self.db.query(Transaction).filter(
            and_(
                Transaction.user_id == user_id,
                Transaction.created_at >= start_datetime,
                Transaction.created_at <= end_datetime
            )
        )
        
        if transaction_type != "all":
            query = query.filter(Transaction.type == transaction_type)
        
        transactions = query.all()
        
        categories: Dict[str, Decimal] = {}
        total = Decimal("0")
        
        for t in transactions:
            amount = t.amount
            if t.type == "expense":
                amount = -amount
            
            if t.category in categories:
                categories[t.category] += amount
            else:
                categories[t.category] = amount
            total += amount
        
        return {
            "type": transaction_type,
            "categories": {k: float(v) for k, v in categories.items()},
            "total": float(total)
        }

    def export_report_excel(self, report_data: dict, report_type: str) -> BytesIO:
        """
        导出报表为Excel格式
        
        Args:
            report_data: 报表数据
            report_type: 报表类型（daily/weekly/monthly/custom）
            
        Returns:
            BytesIO: Excel文件字节流
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            raise ImportError("请安装openpyxl库: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "报表"
        
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        normal_font = Font(size=11)
        
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        if report_type == "daily":
            ws['A1'] = f"每日报表 - {report_data.get('date', '')}"
        elif report_type == "weekly":
            ws['A1'] = f"每周报表 - {report_data.get('start_date', '')} 至 {report_data.get('end_date', '')}"
        elif report_type == "monthly":
            ws['A1'] = f"每月报表 - {report_data.get('year', '')}年{report_data.get('month', '')}月"
        else:
            ws['A1'] = f"自定义报表 - {report_data.get('start_date', '')} 至 {report_data.get('end_date', '')}"
        
        ws['A1'].font = header_font
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "总收入"
        ws['B3'] = f"¥{report_data.get('income', 0):.2f}"
        ws['A4'] = "总支出"
        ws['B4'] = f"¥{report_data.get('expense', 0):.2f}"
        ws['A5'] = "净利润"
        ws['B5'] = f"¥{report_data.get('profit', 0):.2f}"
        ws['A6'] = "订单数量"
        ws['B6'] = report_data.get('order_count', 0)
        ws['A7'] = "订单金额"
        ws['B7'] = f"¥{report_data.get('order_amount', 0):.2f}"
        
        for row in range(3, 8):
            ws[f'A{row}'].font = title_font
            ws[f'B{row}'].font = normal_font
        
        income_categories = report_data.get('income_categories', {})
        if income_categories:
            ws['A9'] = "收入分类明细"
            ws['A9'].font = title_font
            ws.merge_cells('A9:D9')
            
            ws['A10'] = "分类名称"
            ws['B10'] = "金额"
            ws['C10'] = "占比"
            for col in ['A', 'B', 'C']:
                ws[f'{col}10'].fill = header_fill
                ws[f'{col}10'].font = header_font_white
                ws[f'{col}10'].border = thin_border
                ws[f'{col}10'].alignment = center_alignment
            
            total_income = report_data.get('income', 0) or 1
            row = 11
            for category, amount in income_categories.items():
                ws[f'A{row}'] = category
                ws[f'B{row}'] = f"¥{amount:.2f}"
                ws[f'C{row}'] = f"{(amount / total_income * 100):.1f}%"
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = thin_border
                row += 1
        
        expense_categories = report_data.get('expense_categories', {})
        if expense_categories:
            start_row = 11 + len(income_categories) + 2
            ws[f'A{start_row}'] = "支出分类明细"
            ws[f'A{start_row}'].font = title_font
            ws.merge_cells(f'A{start_row}:D{start_row}')
            
            ws[f'A{start_row + 1}'] = "分类名称"
            ws[f'B{start_row + 1}'] = "金额"
            ws[f'C{start_row + 1}'] = "占比"
            for col in ['A', 'B', 'C']:
                ws[f'{col}{start_row + 1}'].fill = header_fill
                ws[f'{col}{start_row + 1}'].font = header_font_white
                ws[f'{col}{start_row + 1}'].border = thin_border
                ws[f'{col}{start_row + 1}'].alignment = center_alignment
            
            total_expense = report_data.get('expense', 0) or 1
            row = start_row + 2
            for category, amount in expense_categories.items():
                ws[f'A{row}'] = category
                ws[f'B{row}'] = f"¥{amount:.2f}"
                ws[f'C{row}'] = f"{(amount / total_expense * 100):.1f}%"
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = thin_border
                row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_report_pdf(self, report_data: dict, report_type: str) -> BytesIO:
        """
        导出报表为PDF格式
        
        Args:
            report_data: 报表数据
            report_type: 报表类型（daily/weekly/monthly/custom）
            
        Returns:
            BytesIO: PDF文件字节流
        """
        html_content = self._generate_report_html(report_data, report_type)
        
        try:
            from weasyprint import HTML
        except ImportError:
            raise ImportError("请安装weasyprint库: pip install weasyprint")
        
        output = BytesIO()
        HTML(string=html_content).write_pdf(output)
        output.seek(0)
        return output

    def _generate_report_html(self, report_data: dict, report_type: str) -> str:
        """
        生成报表HTML内容
        
        Args:
            report_data: 报表数据
            report_type: 报表类型
            
        Returns:
            str: HTML内容
        """
        if report_type == "daily":
            title = f"每日报表 - {report_data.get('date', '')}"
        elif report_type == "weekly":
            title = f"每周报表 - {report_data.get('start_date', '')} 至 {report_data.get('end_date', '')}"
        elif report_type == "monthly":
            title = f"每月报表 - {report_data.get('year', '')}年{report_data.get('month', '')}月"
        else:
            title = f"自定义报表 - {report_data.get('start_date', '')} 至 {report_data.get('end_date', '')}"
        
        income_categories_html = ""
        income_categories = report_data.get('income_categories', {})
        if income_categories:
            total_income = report_data.get('income', 0) or 1
            rows = ""
            for category, amount in income_categories.items():
                rows += f"""
                <tr>
                    <td>{category}</td>
                    <td>¥{amount:.2f}</td>
                    <td>{(amount / total_income * 100):.1f}%</td>
                </tr>
                """
            income_categories_html = f"""
            <h3>收入分类明细</h3>
            <table>
                <thead>
                    <tr>
                        <th>分类名称</th>
                        <th>金额</th>
                        <th>占比</th>
                    </tr>
                </thead>
                <tbody>
                    {rows}
                </tbody>
            </table>
            """
        
        expense_categories_html = ""
        expense_categories = report_data.get('expense_categories', {})
        if expense_categories:
            total_expense = report_data.get('expense', 0) or 1
            rows = ""
            for category, amount in expense_categories.items():
                rows += f"""
                <tr>
                    <td>{category}</td>
                    <td>¥{amount:.2f}</td>
                    <td>{(amount / total_expense * 100):.1f}%</td>
                </tr>
                """
            expense_categories_html = f"""
            <h3>支出分类明细</h3>
            <table>
                <thead>
                    <tr>
                        <th>分类名称</th>
                        <th>金额</th>
                        <th>占比</th>
                    </tr>
                </thead>
                <tbody>
                    {rows}
                </tbody>
            </table>
            """
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>{title}</title>
            <style>
                body {{
                    font-family: 'Microsoft YaHei', Arial, sans-serif;
                    padding: 20px;
                    color: #333;
                }}
                h1 {{
                    color: #3B82F6;
                    text-align: center;
                    border-bottom: 2px solid #3B82F6;
                    padding-bottom: 10px;
                }}
                h3 {{
                    color: #374151;
                    margin-top: 20px;
                }}
                .summary {{
                    display: flex;
                    justify-content: space-around;
                    margin: 20px 0;
                    padding: 15px;
                    background-color: #F9FAFB;
                    border-radius: 8px;
                }}
                .summary-item {{
                    text-align: center;
                }}
                .summary-item .label {{
                    font-size: 14px;
                    color: #6B7280;
                }}
                .summary-item .value {{
                    font-size: 24px;
                    font-weight: bold;
                    margin-top: 5px;
                }}
                .income {{ color: #52C41A; }}
                .expense {{ color: #F5222D; }}
                .profit {{ color: #3B82F6; }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 10px 0;
                }}
                th, td {{
                    border: 1px solid #E5E7EB;
                    padding: 10px;
                    text-align: left;
                }}
                th {{
                    background-color: #3B82F6;
                    color: white;
                }}
                tr:nth-child(even) {{
                    background-color: #F9FAFB;
                }}
            </style>
        </head>
        <body>
            <h1>{title}</h1>
            
            <div class="summary">
                <div class="summary-item">
                    <div class="label">总收入</div>
                    <div class="value income">¥{report_data.get('income', 0):.2f}</div>
                </div>
                <div class="summary-item">
                    <div class="label">总支出</div>
                    <div class="value expense">¥{report_data.get('expense', 0):.2f}</div>
                </div>
                <div class="summary-item">
                    <div class="label">净利润</div>
                    <div class="value profit">¥{report_data.get('profit', 0):.2f}</div>
                </div>
                <div class="summary-item">
                    <div class="label">订单数量</div>
                    <div class="value">{report_data.get('order_count', 0)}</div>
                </div>
                <div class="summary-item">
                    <div class="label">订单金额</div>
                    <div class="value">¥{report_data.get('order_amount', 0):.2f}</div>
                </div>
            </div>
            
            {income_categories_html}
            {expense_categories_html}
            
            <p style="text-align: center; color: #9CA3AF; margin-top: 30px;">
                报表生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
            </p>
        </body>
        </html>
        """
        return html
